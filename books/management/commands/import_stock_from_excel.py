import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from django.core.management.base import BaseCommand, CommandError

from books.models import Book


def normalize_header(value):
    if value is None:
        return ''
    value = str(value).strip().upper()
    value = unicodedata.normalize('NFKD', value)
    return ''.join(ch for ch in value if not unicodedata.combining(ch)).replace(' ', '').replace('_', '')


def find_header_row(worksheet, candidates):
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [normalize_header(cell) for cell in row]
        if all(any(candidate in cell for cell in normalized if cell) for candidate in candidates):
            return row_index, normalized
    return None, None


def resolve_column_index(headers, selection, candidates):
    normalized = [normalize_header(cell) for cell in headers]
    if selection:
        selection_norm = normalize_header(selection)
        for index, value in enumerate(normalized):
            if selection_norm == value:
                return index
    for candidate in candidates:
        for index, value in enumerate(normalized):
            if candidate == value or candidate in value or value in candidate:
                return index
    return None


class Command(BaseCommand):
    help = 'Import stock levels from an Excel stock mapping file.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', help='Path to the XLSX file.')
        parser.add_argument('--sheet', default='B2B stavy OJ', help='Worksheet name to import from.')
        parser.add_argument('--code-column', default='SORT_IDx', help='Column name for product code.')
        parser.add_argument('--quantity-column', default='MNOZSTVI', help='Column name for stock quantity.')
        parser.add_argument('--blocked-column', default='BLOKOVANO', help='Column name for blocked stock quantity.')
        parser.add_argument('--sheet-header-row', type=int, default=0, help='Header row number if automatic detection fails.')

    def handle(self, *args, **options):
        xlsx_path = Path(options['xlsx_path'])
        if not xlsx_path.exists():
            raise CommandError(f'XLSX file not found: {xlsx_path}')

        workbook = load_workbook(xlsx_path, read_only=True)
        sheet_name = options['sheet']
        if sheet_name not in workbook.sheetnames:
            raise CommandError(f'Sheet {sheet_name} not found in workbook. Available sheets: {workbook.sheetnames}')

        worksheet = workbook[sheet_name]
        header_row = options['sheet_header_row']
        normalized_headers = None

        if header_row:
            header = next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            normalized_headers = [normalize_header(cell) for cell in header]
        else:
            header_row, normalized_headers = find_header_row(worksheet, [options['code_column'], options['quantity_column']])
            if header_row is None:
                raise CommandError('Unable to detect header row automatically. Use --sheet-header-row to specify it.')

        code_index = resolve_column_index(normalized_headers, options['code_column'], ['SORT_IDX', 'SORTKOD', 'KODZBOZI', 'ASSORTMENTKEY'])
        quantity_index = resolve_column_index(normalized_headers, options['quantity_column'], ['MNOZSTVI', 'AMOUNT', 'QUANTITY', 'VIRTMNOZ'])
        blocked_index = resolve_column_index(normalized_headers, options['blocked_column'], ['BLOKOVANO', 'BLOCKEDAMOUNT'])

        if code_index is None or quantity_index is None:
            raise CommandError('Could not resolve required columns for code or quantity. Please verify the header names.')

        commands = []
        total_rows = 0
        updated = 0
        skipped = 0
        missing = 0

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(cell is not None for cell in row):
                continue
            total_rows += 1
            code = row[code_index]
            qty = row[quantity_index]
            blocked = row[blocked_index] if blocked_index is not None else None

            if code is None:
                skipped += 1
                continue

            code = str(code).strip()
            try:
                quantity = int(float(qty)) if qty is not None else 0
            except (TypeError, ValueError):
                quantity = 0

            book = Book.find_by_external_code(code)
            if not book:
                missing += 1
                continue

            book.stock = quantity
            book.available = quantity > 0
            book.save()
            updated += 1

        self.stdout.write(self.style.SUCCESS(f'Processed {total_rows} rows, updated {updated} books, skipped {skipped}, missing {missing}.'))
